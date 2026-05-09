#!/usr/bin/env python3
"""
Seed stars_academy.db from StarsAcademy_v2.xlsx.

Reads Players, Attendance Log, Sessions Purchased sheets and creates
all tables + two default users (admin / recorder).

Usage:  python3 scripts/seed.py [path_to_xlsx]
        Defaults to ../StarsAcademy_v2.xlsx relative to this script.
"""

import os, sqlite3, sys
from datetime import datetime
from pathlib import Path

import openpyxl

# ── Paths ────────────────────────────────────────────
SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_DIR = SCRIPT_DIR.parent
DB_PATH = PROJECT_DIR / "stars_academy.db"

XLSX_DEFAULT = PROJECT_DIR.parent / "StarsAcademy_v2.xlsx"
XLSX_PATH = Path(sys.argv[1]) if len(sys.argv) > 1 else XLSX_DEFAULT

if not XLSX_PATH.exists():
    print(f"ERROR: Cannot find {XLSX_PATH}")
    sys.exit(1)

print(f"Source:   {XLSX_PATH}")
print(f"Database: {DB_PATH}")

# ── Database setup ───────────────────────────────────
# Don't delete existing db — preserve users created by the app.
# Instead, clear and re-import data tables only.

conn = sqlite3.connect(str(DB_PATH))
conn.execute("PRAGMA journal_mode=DELETE")
conn.execute("PRAGMA foreign_keys=ON")

# Clear existing data (keep users table intact)
for tbl in ["sessions_purchased", "attendance_log", "players"]:
    try:
        conn.execute(f"DELETE FROM {tbl}")
    except Exception:
        pass

conn.executescript("""
    CREATE TABLE IF NOT EXISTS users (
        id            INTEGER PRIMARY KEY AUTOINCREMENT,
        username      TEXT    NOT NULL UNIQUE,
        password_hash TEXT    NOT NULL,
        role          TEXT    NOT NULL DEFAULT 'recorder'
                      CHECK (role IN ('admin','recorder')),
        created_at    TEXT    NOT NULL DEFAULT (datetime('now'))
    );

    CREATE TABLE IF NOT EXISTS sessions (
        id         INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id    INTEGER NOT NULL REFERENCES users(id),
        token      TEXT    NOT NULL UNIQUE,
        expires_at TEXT    NOT NULL
    );

    CREATE TABLE IF NOT EXISTS players (
        id          INTEGER PRIMARY KEY AUTOINCREMENT,
        code        TEXT    NOT NULL UNIQUE,
        name        TEXT    NOT NULL,
        country     TEXT,
        source      TEXT,
        play_status TEXT    NOT NULL DEFAULT 'Active'
                    CHECK (play_status IN ('Active','Inactive','Left')),
        scholarship INTEGER NOT NULL DEFAULT 0,
        notes       TEXT,
        created_at  TEXT    NOT NULL DEFAULT (datetime('now')),
        updated_at  TEXT    NOT NULL DEFAULT (datetime('now'))
    );

    CREATE TABLE IF NOT EXISTS attendance_log (
        id           INTEGER PRIMARY KEY AUTOINCREMENT,
        player_id    INTEGER NOT NULL REFERENCES players(id),
        session_date TEXT    NOT NULL,
        attended     INTEGER NOT NULL DEFAULT 1,
        created_at   TEXT    NOT NULL DEFAULT (datetime('now'))
    );

    CREATE TABLE IF NOT EXISTS sessions_purchased (
        id                 INTEGER PRIMARY KEY AUTOINCREMENT,
        player_id          INTEGER NOT NULL REFERENCES players(id),
        purchase_date      TEXT    NOT NULL,
        type               TEXT    NOT NULL DEFAULT 'Purchase'
                           CHECK (type IN ('Purchase','Adjustment','Transfer','Opening balance')),
        amount_paid        REAL    NOT NULL DEFAULT 0,
        sessions_purchased INTEGER NOT NULL DEFAULT 0,
        package            TEXT,
        bank_ref           TEXT,
        notes              TEXT,
        created_at         TEXT    NOT NULL DEFAULT (datetime('now'))
    );

    CREATE INDEX IF NOT EXISTS idx_att_player ON attendance_log(player_id);
    CREATE INDEX IF NOT EXISTS idx_att_date   ON attendance_log(session_date);
    CREATE INDEX IF NOT EXISTS idx_sp_player  ON sessions_purchased(player_id);
""")

# Users are created automatically by the app on startup — no need to seed them here.

# ── Load workbook ────────────────────────────────────
wb = openpyxl.load_workbook(str(XLSX_PATH), data_only=True)

# ── Import Players ───────────────────────────────────
ws = wb["Players"]
player_map = {}  # code -> db id

def safe_str(val):
    if val is None:
        return None
    return str(val).strip()

def get_play_status(val):
    s = safe_str(val)
    if s in ("Active", "Inactive", "Left"):
        return s
    return "Active"

row = 2
count_players = 0
while True:
    code = safe_str(ws.cell(row, 1).value)
    name = safe_str(ws.cell(row, 2).value)
    if not code or not name:
        break

    country = safe_str(ws.cell(row, 3).value)
    source = safe_str(ws.cell(row, 4).value)

    # Play Status is col G, Scholarship is col H
    play_status = get_play_status(ws.cell(row, 7).value)
    scholarship_val = ws.cell(row, 8).value
    scholarship = 1 if scholarship_val and str(scholarship_val).strip().lower() in ("scholarship", "yes", "1") else 0

    cursor = conn.execute(
        """INSERT INTO players (code, name, country, source, play_status, scholarship)
           VALUES (?, ?, ?, ?, ?, ?)""",
        (code, name, country, source, play_status, scholarship)
    )
    player_map[code] = cursor.lastrowid
    count_players += 1
    row += 1

print(f"Imported {count_players} players.")

# ── Import Attendance Log ────────────────────────────
ws = wb["Attendance Log"]
count_att = 0

row = 2
while True:
    session_date = ws.cell(row, 1).value
    player_code = safe_str(ws.cell(row, 2).value)
    # Col 3 = Player Name (skip), Col 4 = Attended
    attended_val = ws.cell(row, 4).value

    if session_date is None and player_code is None:
        break
    if session_date is None or player_code is None:
        row += 1
        continue

    if isinstance(session_date, datetime):
        date_str = session_date.strftime("%Y-%m-%d")
    else:
        date_str = str(session_date)

    player_id = player_map.get(player_code)
    if player_id is None:
        row += 1
        continue

    attended = 1 if attended_val is not None and int(attended_val) == 1 else 0

    conn.execute(
        "INSERT INTO attendance_log (player_id, session_date, attended) VALUES (?, ?, ?)",
        (player_id, date_str, attended)
    )
    count_att += 1
    row += 1

print(f"Imported {count_att} attendance records.")

# ── Import Sessions Purchased ────────────────────────
ws = wb["Sessions Purchased"]
count_sp = 0

row = 2
while True:
    purchase_date = ws.cell(row, 1).value
    if purchase_date is None:
        break

    if isinstance(purchase_date, datetime):
        date_str = purchase_date.strftime("%Y-%m-%d")
    else:
        date_str = str(purchase_date)

    sp_type      = safe_str(ws.cell(row, 2).value) or "Purchase"
    player_code  = safe_str(ws.cell(row, 3).value)
    # col 4 = player name (skip, we have it from players table)
    amount_paid  = ws.cell(row, 5).value or 0
    sessions     = ws.cell(row, 6).value or 0
    package      = safe_str(ws.cell(row, 7).value)
    bank_ref     = safe_str(ws.cell(row, 8).value)
    notes        = safe_str(ws.cell(row, 9).value)

    # Validate type
    valid_types = ("Purchase", "Adjustment", "Transfer", "Opening balance")
    if sp_type not in valid_types:
        sp_type = "Purchase"

    player_id = player_map.get(player_code)
    if player_id is None:
        row += 1
        continue

    conn.execute(
        """INSERT INTO sessions_purchased
           (player_id, purchase_date, type, amount_paid, sessions_purchased, package, bank_ref, notes)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
        (player_id, date_str, sp_type, float(amount_paid), int(sessions), package, bank_ref, notes)
    )
    count_sp += 1
    row += 1

print(f"Imported {count_sp} session purchase records.")

# ── Commit & verify ─────────────────────────────────
conn.commit()

# Quick verification
player_count = conn.execute("SELECT COUNT(*) FROM players").fetchone()[0]
att_count = conn.execute("SELECT COUNT(*) FROM attendance_log").fetchone()[0]
sp_count = conn.execute("SELECT COUNT(*) FROM sessions_purchased").fetchone()[0]

print(f"\nVerification:")
print(f"  Players:            {player_count}")
print(f"  Attendance records: {att_count}")
print(f"  Purchase records:   {sp_count}")

# Show a few dashboard rows as sanity check
print(f"\nSample dashboard (first 5 players):")
rows = conn.execute("""
    SELECT
        p.code, p.name,
        COALESCE(sp.total, 0) AS paid,
        COALESCE(att.total, 0) AS attended,
        COALESCE(sp.total, 0) - COALESCE(att.total, 0) AS balance
    FROM players p
    LEFT JOIN (SELECT player_id, SUM(sessions_purchased) AS total FROM sessions_purchased GROUP BY player_id) sp
        ON sp.player_id = p.id
    LEFT JOIN (SELECT player_id, COUNT(*) AS total FROM attendance_log WHERE attended=1 GROUP BY player_id) att
        ON att.player_id = p.id
    ORDER BY p.code LIMIT 5
""").fetchall()

for r in rows:
    print(f"  {r[0]} {r[1]:<15} paid={r[2]:<4} attended={r[3]:<4} balance={r[4]}")

conn.close()
print(f"\nDone. Database saved to {DB_PATH}")
